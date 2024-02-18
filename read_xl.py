import requests
from bs4 import BeautifulSoup
import time
import re
import locale
import math
import random
import openpyxl


p_user = 'rjjtzjxw'
p_pass = 'u8aovf7qutdn'
ips = ['23.109.208.30:6554', '104.238.8.190:6048', '23.109.225.254:5885', '104.238.8.100:5958', '23.109.232.242:6162',
       '154.92.125.62:5363', '104.238.9.112:6565', '154.92.124.232:5260', '104.238.9.90:6543', '23.109.225.56:5687',
       '104.222.187.196:6320', '23.109.225.185:5816', '23.109.225.168:5799', '23.109.225.93:5724', '154.92.124.80:5108',
       '23.109.208.183:6707', '104.222.187.248:6372', '23.109.232.79:5999', '154.92.124.121:5149', '154.92.124.7:5035',
       '104.222.187.217:6341', '104.222.187.89:6213', '23.109.232.237:6157', '23.109.232.93:6013',
       '104.222.187.182:6306', '104.238.9.198:6651', '104.222.185.235:5798', '109.207.130.173:8180',
       '23.109.208.108:6632', '23.109.232.41:5961', '154.92.125.198:5499', '154.92.125.215:5516',
       '109.207.130.222:8229', '104.238.8.9:5867', '23.109.232.166:6086', '109.207.130.45:8052', '23.109.208.81:6605',
       '154.92.124.92:5120', '104.238.9.50:6503', '23.109.225.166:5797', '104.222.185.119:5682', '23.109.232.103:6023',
       '23.109.208.147:6671', '154.92.125.119:5420', '23.109.225.124:5755', '104.238.8.37:5895', '104.238.8.164:6022',
       '104.222.185.244:5807', '154.92.124.77:5105', '104.238.8.109:5967', '104.238.9.57:6510', '154.92.124.85:5113',
       '109.207.130.5:8012', '154.92.125.97:5398', '104.222.185.209:5772', '23.109.219.56:6280', '23.109.208.28:6552',
       '154.92.124.159:5187', '104.238.9.108:6561', '23.109.208.27:6551', '23.109.219.90:6314', '104.222.187.113:6237',
       '23.109.225.117:5748', '23.109.208.75:6599', '104.238.8.24:5882', '23.109.208.38:6562', '109.207.130.171:8178',
       '109.207.130.32:8039', '23.109.219.39:6263', '104.238.8.7:5865', '104.222.187.17:6141', '154.92.124.202:5230',
       '104.222.187.100:6224', '109.207.130.89:8096', '104.222.185.237:5800', '104.238.8.44:5902', '23.109.219.97:6321',
       '23.109.225.147:5778', '104.222.185.172:5735', '104.238.8.110:5968', '104.222.187.18:6142',
       '23.109.232.101:6021', '23.109.225.182:5813', '154.92.125.195:5496', '23.109.232.232:6152',
       '104.222.187.114:6238', '104.222.185.4:5567', '23.109.219.164:6388', '104.222.187.119:6243',
       '104.238.9.110:6563', '23.109.208.46:6570', '154.92.125.206:5507', '154.92.125.102:5403', '23.109.225.4:5635',
       '104.222.187.48:6172', '109.207.130.176:8183', '154.92.124.96:5124', '104.238.8.81:5939', '23.109.225.140:5771',
       '104.238.9.166:6619']


def switch_proxy(ses):
    global p_pass
    global p_user
    global ips
    print('switching proxy')
    proxies = {
                'http': f"http://{p_user}:{p_pass}@{ips[random.randint(0, len(ips)-1)]}"
        }
    ses.proxies = proxies
    

def extract_ids_from_excel(file_path, sheet_name, start_cell, end_cell):
    ids = []
    # Load the workbook
    wb = openpyxl.load_workbook(file_path)
    # Select the active sheet
    sheet = wb[sheet_name]
    
    # Iterate through the specified range of cells
    for row in sheet[start_cell:end_cell]:
        for cell in row:
            # Assuming each cell contains a single ID
            ids.append(cell.value)
    
    return ids

def write_to_specific_cell(workbook, cell, data):
    # Get the active sheet
    ws = workbook.active
    
    # Write data to the specified cell
    ws[cell] = data


# Example usage
file_path = "resource_data_new.xlsx"
sheet_name = "Sheet"  # Change to your sheet name
start_cell = "G2"
end_cell = "G1693"
ids = extract_ids_from_excel(file_path, sheet_name, start_cell, end_cell)

spy_data = {'spy': 'שלח את המרגלים'}
login_url = "http://s1.izra.co.il/login"

def getP(session, page):
    while True:
        try:
            response = session.get(page, timeout = 1)
            if response.status_code == 200:
                break
            else:
                switch_proxy(session)
        except:
            continue
    return response

def login():
    proxies = {
        'http': f"http://{p_user}:{p_pass}@{ips[random.randint(0, len(ips)-1)]}"
    }
    s = requests.Session()
    s.proxies = proxies
    custom_user_agent = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/115.0.0.0 Safari/537.36"
    s.headers.update({'User-Agent': custom_user_agent})
    while True:
        try:
            response = s.post(login_url, data = login_data, timeout = 1)
            if response.status_code == 200:
                break
            else:
                switch_proxy(s)
        except:
            continue
        time.sleep(0.5)
    return s

def scout(session, url, spy_data):
    while True:
        try:
            response = session.post(url, data = spy_data, timeout = 1)
            if response.status_code == 200:
                break
            else:
                switch_proxy(session)
        except:
            continue
    return response

def format_with_commas(number):
    locale.setlocale(locale.LC_ALL, '')
    return locale.format_string("%d", number, grouping=True)
login_data = {"email": 'shaharizra1@gmail.com', "password": '1234512345', "rem": "on", "reg": "התחברות >>"}


s = login()
try:
    wb = openpyxl.load_workbook(file_path)
except FileNotFoundError:
    wb = openpyxl.load_workbook(file_path)


for index, number in enumerate(ids):
    print(index)
    response = scout(s, f"http://s1.izra.co.il/spy/spyuser/?targetId={number}", spy_data)
    soup = BeautifulSoup(response.content, 'html.parser')
    span_tag = soup.find('span', class_='row color-gold')
    if 'חופשות שנוצלו' in soup.text:
        summ = 'vacation'
    elif 'war_no_rcity' in soup.text:
        summ = 'diff city'
    elif span_tag is not None:
        gold = int(span_tag.get_text().replace(',', ''))
        span_tag = soup.find('span', class_='row color-iron')
        iron = int(span_tag.get_text().replace(',', ''))
        span_tag = soup.find('span', class_='row color-wood')
        wood = int(span_tag.get_text().replace(',', ''))
        span_tag = soup.find('span', class_='row color-food')
        food = int(span_tag.get_text().replace(',', ''))
        slaves = int(re.findall('<span class="row">עבדים</span>\n<span class="row color-food">([0-9,]*)</span>', str(soup))[0].replace(',',''))
        farmers = int(re.findall('<span class="row">איכרים</span>\n<span class="row color-food">([0-9,]*)</span>', str(soup))[0].replace(',',''))
        summ = food + gold + iron + wood
        write_to_specific_cell(wb, f'C{index+2}', gold)
        write_to_specific_cell(wb, f'D{index+2}', iron)
        write_to_specific_cell(wb, f'E{index+2}', wood)
        write_to_specific_cell(wb, f'F{index+2}', food)
        write_to_specific_cell(wb, f'I{index+2}', slaves)
    write_to_specific_cell(wb, f'B{index+2}', summ)





wb.save(file_path)
 

